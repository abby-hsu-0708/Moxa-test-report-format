import streamlit as st
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
import os
import io
import re
import html

# 設定網頁標題與外觀
st.set_page_config(
    page_title="SN 資料自動排序與分組格式化工具",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded"
)

# 注入自訂 CSS 樣式，提升視覺美感與專業感
st.markdown("""
<style>
    /* 引入現代字型 */
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&family=Outfit:wght@300;400;500;600;700&display=swap');
    
    html, body, [class*="css"] {
        font-family: 'Inter', 'Outfit', -apple-system, BlinkMacSystemFont, sans-serif;
    }
    
    /* 漸層標題 */
    .main-title {
        font-size: 2.8rem;
        font-weight: 700;
        background: linear-gradient(135deg, #00C6FF 0%, #0072FF 100%);
        -webkit-background-clip: text;
        -webkit-text-fill-color: transparent;
        margin-bottom: 0.5rem;
    }
    
    .subtitle {
        font-size: 1.1rem;
        color: #64748B;
        margin-bottom: 2rem;
    }
    
    /* 資訊卡片 */
    .metric-card {
        background: rgba(255, 255, 255, 0.8);
        border: 1px solid #E2E8F0;
        border-radius: 12px;
        padding: 1.5rem;
        box-shadow: 0 4px 6px -1px rgba(0, 0, 0, 0.05), 0 2px 4px -1px rgba(0, 0, 0, 0.03);
        transition: transform 0.2s, box-shadow 0.2s;
    }
    
    .metric-card:hover {
        transform: translateY(-2px);
        box-shadow: 0 10px 15px -3px rgba(0, 0, 0, 0.1), 0 4px 6px -2px rgba(0, 0, 0, 0.05);
    }
    
    .metric-title {
        font-size: 0.875rem;
        font-weight: 600;
        color: #64748B;
        text-transform: uppercase;
        letter-spacing: 0.05em;
    }
    
    .metric-value {
        font-size: 1.875rem;
        font-weight: 700;
        color: #0F172A;
        margin-top: 0.25rem;
    }
</style>
""", unsafe_allow_html=True)

# 顯示漂亮的主頁面標題
st.markdown("<h1 class='main-title'>📊 SN 資料排序與分組格式化系統</h1>", unsafe_allow_html=True)
st.markdown("<p class='subtitle'>依據 Model Name 排序，以同個 Model 且相同的 Ship_NO、ShippingDate 為組別拆分，SN 最多一列顯示 5 個，自動生成符合 Panny 格式之 Excel 報表。</p>", unsafe_allow_html=True)

# 偵測本地是否有預設的 Panny.xlsx 樣板
default_template_exists = os.path.exists("Panny.xlsx")

# 側邊欄設定
with st.sidebar:
    st.markdown("### ⚙️ 系統設定")
    
    # 樣板來源選擇
    st.markdown("#### 樣板 Excel 來源")
    template_option = st.radio(
        "選擇 Template 來源：",
        ["使用本地 Panny.xlsx" if default_template_exists else "本地無 Panny.xlsx (自動生成格式)", "上傳自訂 Panny 樣板"],
        index=0
    )
    
    uploaded_template = None
    if "上傳自訂" in template_option:
        uploaded_template = st.file_uploader("請上傳 Panny.xlsx 樣板檔案", type=["xlsx"])
        
    st.markdown("---")
    st.markdown("### 📋 格式規範說明")
    st.info("""
    1. **排序邏輯**：資料將依據 Model Name (即 Part_DESC) 排序。
    2. **換列條件**：同一個 Model 若出貨單號 (Ship_NO) 或出貨日期 (ShippingDate) 不同，則分拆至另一列。
    3. **SN 欄位上限**：同一列最多填入 5 個 SN (顯示於 Q 至 U 欄)。若超出 5 個，則將剩餘的 SN 分拆至下一列展示（左側所有欄位資訊均會重複填入，以利後續資料分析）。
    """)

# 🔍 步驟零：電子郵件 SN 自動提取整理
st.subheader("🔍 步驟零：電子郵件 SN 自動提取整理")

# 使用全前端 HTML/JS 實作，以解決 Streamlit 必須按 Ctrl+Enter 才能生效的技術限制。
# 這樣一來，使用者不論打字或貼上文字，下方都會在一瞬間即時顯示整理後的 SN，不打斷中文輸入法。
step_zero_html = r"""
<!DOCTYPE html>
<html>
<head>
    <meta charset="utf-8">
    <style>
        @import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&family=Outfit:wght@300;400;500;600;700&display=swap');
        
        body {
            font-family: 'Inter', 'Outfit', -apple-system, BlinkMacSystemFont, sans-serif;
            background-color: transparent;
            margin: 0;
            padding: 0;
            color: #0F172A;
            overflow: hidden;
        }
        
        .info-card {
            background-color: #F8FAFC; 
            border: 1px solid #E2E8F0; 
            border-radius: 12px; 
            padding: 1.2rem; 
            margin-bottom: 1rem;
        }
        
        .info-text {
            font-weight: 600; 
            color: #0F172A; 
            margin: 0 0 0.5rem 0; 
            font-size: 0.95rem;
        }
        
        .label-text {
            font-size: 14px; 
            font-weight: 500; 
            color: #334155; 
            display: block; 
            margin-bottom: 8px;
        }
        
        .text-area {
            width: 100%;
            height: 180px;
            padding: 12px;
            border-radius: 8px;
            border: 1px solid #CBD5E1;
            font-family: inherit;
            font-size: 14px;
            box-sizing: border-box;
            resize: none;
            outline: none;
            transition: border-color 0.2s, box-shadow 0.2s;
        }
        
        .text-area:focus {
            border-color: #0072FF;
            box-shadow: 0 0 0 3px rgba(0, 114, 255, 0.15);
        }
        
        .button-row {
            margin-top: 10px;
            margin-bottom: 20px;
            display: flex;
            gap: 12px;
        }
        
        .btn-clear {
            background-color: #F1F5F9;
            color: #475569;
            border: 1px solid #E2E8F0;
            padding: 8px 16px;
            border-radius: 6px;
            cursor: pointer;
            font-weight: 600;
            font-size: 13px;
            transition: all 0.2s ease;
            display: flex;
            align-items: center;
        }
        
        .btn-clear:hover {
            background-color: #E2E8F0;
            color: #1E293B;
        }
        
        .result-title {
            margin: 0 0 8px 0; 
            font-size: 14px; 
            font-weight: 600; 
            color: #334155;
        }
        
        .result-box {
            background-color: #F8FAFC;
            border: 1px solid #E2E8F0;
            border-radius: 6px;
            padding: 12px;
            font-family: monospace;
            font-size: 14px;
            white-space: pre-wrap;
            word-break: break-all;
            margin: 0 0 12px 0;
            height: 180px;
            overflow-y: auto;
            box-sizing: border-box;
        }
        
        .btn-copy {
            background: linear-gradient(135deg, #00C6FF 0%, #0072FF 100%);
            color: white;
            border: none;
            padding: 10px 22px;
            border-radius: 8px;
            cursor: pointer;
            font-weight: 600;
            font-size: 14px;
            box-shadow: 0 4px 6px -1px rgba(0, 0, 0, 0.1);
            transition: all 0.2s ease;
            display: flex;
            align-items: center;
        }
        
        .btn-copy:hover {
            opacity: 0.9;
            transform: translateY(-1px);
        }
        
        .btn-copy:active {
            transform: translateY(0);
        }
        
        .toast {
            color: #10B981; 
            font-weight: 600; 
            font-size: 14px;
            align-self: center;
            display: none;
        }
        
        .warning-card {
            background-color: #FFFBEB; 
            border: 1px solid #FDE68A; 
            border-radius: 8px; 
            padding: 12px; 
            color: #B45309; 
            font-size: 14px;
            display: none;
        }
    </style>
</head>
<body>
    <div class="info-card">
        <p class="info-text">請在此貼上包含多個 Model 的電子郵件或雜訊文字。系統會自動過濾 Model 名稱與雜訊，僅提取出純 SN：</p>
    </div>
    
    <label class="label-text">貼上電子郵件或雜亂文字：</label>
    <textarea id="raw-input" class="text-area" placeholder="例如：
MODEL : EDS-518E-4GTXSFP
TBFED1009476
TBFED1009482

MODEL : IKS-G6524A-4GTXSFP
TBFAD1056579
..."></textarea>
    
    <div class="button-row">
        <button id="clear-btn" class="btn-clear">🗑️ 清除已輸入文字</button>
    </div>
    
    <div id="no-sn-warning" class="warning-card">
        ⚠️ 未在輸入內容中偵測到符合格式的 SN。請確認是否包含純英數且長度介於 8 至 15 碼。
    </div>
    
    <div id="result-section" style="display: none; margin-top: 15px;">
        <h5 class="result-title">✨ 整理後的 SN 列表</h5>
        <div id="result-output" class="result-box"></div>
        
        <div style="display: flex; align-items: center; gap: 12px;">
            <button id="copy-btn" class="btn-copy">📋 複製整理後的 SN 列表</button>
            <span id="copy-success-msg" class="toast">✓ 已複製到剪貼簿！</span>
        </div>
    </div>
    
    <script>
        const rawInput = document.getElementById('raw-input');
        const clearBtn = document.getElementById('clear-btn');
        const resultSection = document.getElementById('result-section');
        const resultOutput = document.getElementById('result-output');
        const copyBtn = document.getElementById('copy-btn');
        const copySuccessMsg = document.getElementById('copy-success-msg');
        const noSnWarning = document.getElementById('no-sn-warning');
        
        // 即時輸入處理事件
        rawInput.addEventListener('input', processInput);
        
        function processInput() {
            const text = rawInput.value;
            if (!text.trim()) {
                resultSection.style.display = 'none';
                noSnWarning.style.display = 'none';
                return;
            }
            
            // 使用正則表達式全局搜索 8-15 碼的純英數字 (排除含減號的 Model 名稱)
            const matches = text.match(/\b[a-zA-Z0-9]{8,15}\b/g) || [];
            
            if (matches.length > 0) {
                const cleanedText = matches.join('\n');
                resultOutput.textContent = cleanedText;
                resultSection.style.display = 'block';
                noSnWarning.style.display = 'none';
            } else {
                resultSection.style.display = 'none';
                noSnWarning.style.display = 'block';
            }
        }
        
        // 清除按鈕事件
        clearBtn.addEventListener('click', () => {
            rawInput.value = '';
            resultSection.style.display = 'none';
            noSnWarning.style.display = 'none';
        });
        
        // 複製按鈕事件
        copyBtn.addEventListener('click', () => {
            const copyText = resultOutput.textContent;
            if (navigator.clipboard && window.isSecureContext) {
                navigator.clipboard.writeText(copyText).then(showToast).catch(fallbackCopy);
            } else {
                fallbackCopy();
            }
        });
        
        function fallbackCopy() {
            const tempTextArea = document.createElement("textarea");
            tempTextArea.value = resultOutput.textContent;
            document.body.appendChild(tempTextArea);
            tempTextArea.select();
            try {
                document.execCommand("copy");
                showToast();
            } catch (err) {
                alert("無法複製，請手動選取結果框中的 SN 進行複製。");
            }
            document.body.removeChild(tempTextArea);
        }
        
        function showToast() {
            copySuccessMsg.style.display = 'inline';
            setTimeout(() => {
                copySuccessMsg.style.display = 'none';
            }, 2000);
        }
    </script>
</body>
</html>
"""

st.components.v1.html(step_zero_html, height=530, scrolling=False)

st.markdown("---")

# 主內容區：上傳 SN Data 檔案
st.subheader("📥 步驟一：上傳 SN Data 檔案")
uploaded_data_file = st.file_uploader("請上傳要處理的 SN Data Excel 檔案 (例如 SN_Data (5).xlsx)", type=["xlsx"])

def process_excel(data_bytes, template_bytes=None):
    """
    資料處理核心邏輯：
    1. 讀取輸入 Excel 中 'SN Data' 頁籤的資料。
    2. 排序與分組處理。
    3. 寫入 openpyxl Template 中，套用格式。
    """
    # 讀取輸入資料
    # 先讀取所有 Sheet，若有 'SN Data' 則讀取之，否則讀取第一個 Sheet
    xls = pd.ExcelFile(io.BytesIO(data_bytes))
    sheet_name = "SN Data" if "SN Data" in xls.sheet_names else xls.sheet_names[0]
    
    df = pd.read_excel(io.BytesIO(data_bytes), sheet_name=sheet_name)
    
    # 先對現有 columns 進行清理與去空白
    df.columns = df.columns.astype(str).str.strip()
    
    # 健壯的 Header 識別：如果 'SN' 或是 '序號' 等關鍵字不在 Columns 中，但第一列有這些欄位
    columns_upper = [c.upper() for c in df.columns]
    has_header_keywords = any(kw in columns_upper for kw in ['SN', 'PART_NO', '序號', '料號'])
    if not has_header_keywords and len(df) > 0:
        first_row_vals = df.iloc[0].astype(str).str.strip().tolist()
        first_row_vals_upper = [v.upper() for v in first_row_vals]
        if any(kw in first_row_vals_upper for kw in ['SN', 'PART_NO', '序號', '料號']):
            # 使用第一列作為新的 Columns，並去除空白
            df.columns = first_row_vals
            df = df[1:].reset_index(drop=True)
            
    # 再次清理與去空白
    df.columns = df.columns.astype(str).str.strip()
    
    # 建立一個大小寫與空白無關的對應 Dictionary
    column_mapping = {}
    for col in df.columns:
        column_mapping[col.upper()] = col
        
    # 定義每個標準欄位的中文與英文別名列表 (全部轉為大寫並去除空白匹配)
    aliases = {
        'SN': ['SN', '序號'],
        'Part_NO': ['PART_NO', '料號'],
        'Part_DESC': ['PART_DESC', 'MODEL', 'MODEL_NAME', 'MODEL NAME', '品名'],
        'Ship_NO': ['SHIP_NO', '出貨單號'],
        'Customer': ['CUSTOMER', '客戶'],
        'ShippingDate': ['SHIPPINGDATE', 'SHIPPING_DATE', '出貨日期', '日期']
    }
    
    # 開始進行同義字對齊與檢查
    rename_dict = {}
    missing_cols = []
    
    for std_name, alias_list in aliases.items():
        found_col = None
        for alias in alias_list:
            alias_upper = alias.upper()
            if alias_upper in column_mapping:
                found_col = column_mapping[alias_upper]
                break
        if found_col is not None:
            rename_dict[found_col] = std_name
        else:
            missing_cols.append(std_name)
            
    if missing_cols:
        raise ValueError(f"輸入檔案缺少以下必要欄位（支援中英文對照）: {', '.join(missing_cols)}")
        
    # 重新命名為標準名稱，以防中英文混用或大小寫不一致導致後續報錯
    df = df.rename(columns=rename_dict)
        
    # 處理出貨日期轉換為字串排序與分組
    if pd.api.types.is_datetime64_any_dtype(df['ShippingDate']):
        df['ShippingDate_Str'] = df['ShippingDate'].dt.strftime('%Y/%m/%d')
    else:
        df['ShippingDate_Str'] = df['ShippingDate'].astype(str).str.replace(r'\s+00:00:00', '', regex=True).str.strip()
        
    # 處理其他可能為 NaN 的欄位，確保分組不會出錯
    df['Part_NO'] = df['Part_NO'].fillna("").astype(str)
    df['Part_DESC'] = df['Part_DESC'].fillna("").astype(str)
    df['Ship_NO'] = df['Ship_NO'].fillna("").astype(str)
    df['Customer'] = df['Customer'].fillna("").astype(str)
    df['SN'] = df['SN'].fillna("").astype(str)
    
    # 排序：依據 Part_DESC (Model Name) -> Ship_NO -> ShippingDate -> SN
    df_sorted = df.sort_values(by=['Part_DESC', 'Ship_NO', 'ShippingDate_Str', 'SN']).copy()
    
    # 初始化 openpyxl 工作簿
    if template_bytes is not None:
        wb = openpyxl.load_workbook(io.BytesIO(template_bytes))
    elif default_template_exists and "上傳自訂" not in template_option:
        wb = openpyxl.load_workbook("Panny.xlsx")
    else:
        wb = openpyxl.Workbook()
        # 移除預設的 Sheet
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])
            
    # 確認 Template 頁籤是否存在，若不存在則建立並初始化格式
    if "Template" not in wb.sheetnames:
        ws = wb.create_sheet("Template")
        
        # 設定欄寬
        col_widths = {
            'A': 7.875, 'B': 3.875, 'C': 13.0, 'D': 13.0, 'E': 13.0, 'F': 13.0, 'G': 13.0,
            'H': 13.0, 'I': 13.0, 'J': 13.0, 'K': 13.0, 'L': 15.625, 'M': 43.125, 'N': 11.125,
            'O': 41.625, 'P': 15.375, 'Q': 13.625, 'R': 13.0, 'S': 13.0, 'T': 15.875, 'U': 13.0,
            'V': 8.875, 'W': 13.0, 'X': 13.0, 'Y': 13.0, 'Z': 13.0, 'AA': 13.0
        }
        for col_letter, width in col_widths.items():
            ws.column_dimensions[col_letter].width = width
            
        # 寫入第 5 行 Header
        headers = {
            1: "SN", 8: "SN1", 9: "SN2", 12: "Part_NO", 13: "Part_DESC", 
            14: "Ship_NO", 15: "Customer", 16: "ShippingDate", 17: "SN"
        }
        
        cyan_fill = PatternFill(start_color='FF00FFFF', end_color='FF00FFFF', fill_type='solid')
        header_font = Font(name='Malgun Gothic Semilight', size=10, bold=False)
        header_align = Alignment(horizontal='left', vertical='center')
        thin_side = Side(style='thin', color='000000')
        header_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
        
        for col_idx, val in headers.items():
            cell = ws.cell(row=5, column=col_idx, value=val)
            cell.font = header_font
            cell.alignment = header_align
            cell.border = header_border
            # L 到 P 欄以及 A 欄有背景色
            if col_idx in [1, 12, 13, 14, 15, 16]:
                cell.fill = cyan_fill
            else:
                cell.fill = PatternFill(start_color='FFFFFFFF', end_color='FFFFFFFF', fill_type='solid')
    else:
        ws = wb["Template"]
        # 清空第 6 列以後的 L 到 U 欄 (Col 12 到 21) 的值與樣式
        max_r = ws.max_row
        if max_r >= 6:
            for r in range(6, max_r + 1):
                for c in range(12, 22):
                    cell = ws.cell(row=r, column=c)
                    cell.value = None
                    cell.font = Font(name='Calibri', size=11)
                    cell.fill = PatternFill(fill_type=None)
                    cell.border = Border()
                    cell.alignment = Alignment()
                    
    # 設定資料格樣式
    thin_side = Side(style='thin', color='000000')
    cell_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    cell_font = Font(name='Malgun Gothic Semilight', size=10, bold=False)
    cell_align = Alignment(horizontal='left', vertical='center')
    cell_fill = PatternFill(start_color='FFFFFFFF', end_color='FFFFFFFF', fill_type='solid')
    
    # 進行分組與寫入
    current_row = 6
    preview_rows = []
    
    # groupby 順序依 sort 排序的結果進行
    # 使用 sort=False 保留 dataframe 本身的排序順序
    grouped = df_sorted.groupby(['Part_NO', 'Part_DESC', 'Ship_NO', 'Customer', 'ShippingDate_Str'], sort=False)
    
    for group_key, group_df in grouped:
        part_no, part_desc, ship_no, customer, shipping_date_str = group_key
        sns = group_df['SN'].tolist()
        
        # 每 5 個 SN 一組拆分
        for i in range(0, len(sns), 5):
            chunk_sns = sns[i:i+5]
            
            # 只有在第一列 (i == 0) 時才寫入左側資訊，後續拆分列則保留空白
            if i == 0:
                ws.cell(row=current_row, column=12, value=part_no)
                ws.cell(row=current_row, column=13, value=part_desc)
                ws.cell(row=current_row, column=14, value=ship_no)
                ws.cell(row=current_row, column=15, value=customer)
                ws.cell(row=current_row, column=16, value=shipping_date_str)
                
                prev_part_no = part_no
                prev_part_desc = part_desc
                prev_ship_no = ship_no
                prev_customer = customer
                prev_shipping_date = shipping_date_str
            else:
                ws.cell(row=current_row, column=12, value=None)
                ws.cell(row=current_row, column=13, value=None)
                ws.cell(row=current_row, column=14, value=None)
                ws.cell(row=current_row, column=15, value=None)
                ws.cell(row=current_row, column=16, value=None)
                
                prev_part_no = ""
                prev_part_desc = ""
                prev_ship_no = ""
                prev_customer = ""
                prev_shipping_date = ""
            
            # 寫入 Q 到 U 欄 (Col 17 到 21)
            for sn_idx in range(5):
                col_idx = 17 + sn_idx
                if sn_idx < len(chunk_sns):
                     ws.cell(row=current_row, column=col_idx, value=chunk_sns[sn_idx])
                else:
                     ws.cell(row=current_row, column=col_idx, value=None)
            
            # 套用樣式
            for c in range(12, 22):
                cell = ws.cell(row=current_row, column=c)
                cell.font = cell_font
                cell.border = cell_border
                cell.alignment = cell_align
                cell.fill = cell_fill
            
            # 記錄預覽資料 (配合 Excel 呈現，非首列亦留空)
            preview_row = {
                "Part_NO": prev_part_no,
                "Part_DESC": prev_part_desc,
                "Ship_NO": prev_ship_no,
                "Customer": prev_customer,
                "ShippingDate": prev_shipping_date,
                "SN1": chunk_sns[0] if len(chunk_sns) > 0 else "",
                "SN2": chunk_sns[1] if len(chunk_sns) > 1 else "",
                "SN3": chunk_sns[2] if len(chunk_sns) > 2 else "",
                "SN4": chunk_sns[3] if len(chunk_sns) > 3 else "",
                "SN5": chunk_sns[4] if len(chunk_sns) > 4 else ""
            }
            preview_rows.append(preview_row)
            
            current_row += 1
            
    # 強制開啟網格線與套用框線
    ws.views.sheetView[0].showGridLines = True
    
    # 將結果儲存於 memory 中以供下載
    out_io = io.BytesIO()
    wb.save(out_io)
    out_io.seek(0)
    
    return out_io.getvalue(), pd.DataFrame(preview_rows), df_sorted.shape[0], len(grouped)

if uploaded_data_file is not None:
    try:
        # 讀取檔案 bytes
        data_bytes = uploaded_data_file.read()
        
        # 處理 Template bytes
        template_bytes = None
        if "上傳自訂" in template_option and uploaded_template is not None:
            template_bytes = uploaded_template.read()
            
        # 執行轉換
        output_excel, preview_df, total_sns, total_groups = process_excel(data_bytes, template_bytes)
        
        st.success("🎉 資料轉換成功！")
        
        # 顯示統計數據卡片
        col1, col2, col3 = st.columns(3)
        with col1:
            st.markdown(f"""
            <div class="metric-card">
                <div class="metric-title">總序號數量 (SN)</div>
                <div class="metric-value">{total_sns} 筆</div>
            </div>
            """, unsafe_allow_html=True)
        with col2:
            st.markdown(f"""
            <div class="metric-card">
                <div class="metric-title">分組總數 (Model+Ship+Date)</div>
                <div class="metric-value">{total_groups} 組</div>
            </div>
            """, unsafe_allow_html=True)
        with col3:
            st.markdown(f"""
            <div class="metric-card">
                <div class="metric-title">格式化後總列數</div>
                <div class="metric-value">{preview_df.shape[0]} 列</div>
            </div>
            """, unsafe_allow_html=True)
            
        st.markdown("<br>", unsafe_allow_html=True)
        
        # 提供下載按鈕
        st.subheader("📥 步驟二：下載產出 Excel 檔案")
        st.download_button(
            label="💾 下載格式化 Excel 報表",
            data=output_excel,
            file_name="Formatted_Panny_Report.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        
        # 顯示資料預覽
        st.subheader("👀 步驟三：格式化結果即時預覽與複製")
        
        # 產生 HTML 格式表格以保留樣式（框線、Header顏色、字型）
        def generate_html_table(df):
            html = []
            html.append('<table style="border-collapse: collapse; font-family: \'Malgun Gothic Semilight\', sans-serif; font-size: 10pt; border: 1px solid #000000;">')
            
            # Header 列樣式 (Row 5)
            html.append('  <tr>')
            cyan_headers = ["Part_NO", "Part_DESC", "Ship_NO", "Customer", "ShippingDate"]
            for h in cyan_headers:
                html.append(f'    <th style="border: 1px solid #000000; background-color: #00FFFF; font-weight: normal; text-align: left; padding: 4px 8px; font-family: \'Malgun Gothic Semilight\';">{h}</th>')
            html.append('    <th style="border: 1px solid #000000; background-color: #FFFFFF; font-weight: normal; text-align: left; padding: 4px 8px; font-family: \'Malgun Gothic Semilight\';">SN</th>')
            for _ in range(4):
                html.append('    <th style="border: 1px solid #000000; background-color: #FFFFFF; font-weight: normal; text-align: left; padding: 4px 8px;"></th>')
            html.append('  </tr>')
            
            # 資料列樣式 (Row 6 之後)
            for _, row in df.iterrows():
                html.append('  <tr>')
                for col in ["Part_NO", "Part_DESC", "Ship_NO", "Customer", "ShippingDate"]:
                    val = str(row[col]) if pd.notna(row[col]) and row[col] != "" else ""
                    html.append(f'    <td style="border: 1px solid #000000; background-color: #FFFFFF; padding: 4px 8px; text-align: left; font-family: \'Malgun Gothic Semilight\';">{val}</td>')
                for col in ["SN1", "SN2", "SN3", "SN4", "SN5"]:
                    val = str(row[col]) if pd.notna(row[col]) and row[col] != "" else ""
                    html.append(f'    <td style="border: 1px solid #000000; background-color: #FFFFFF; padding: 4px 8px; text-align: left; font-family: \'Malgun Gothic Semilight\';">{val}</td>')
                html.append('  </tr>')
                
            html.append('</table>')
            return "\n".join(html)

        html_table_str = generate_html_table(preview_df)
        tsv_str = preview_df.to_csv(index=False, sep='\t')
        
        # 進行 JS 轉義
        escaped_html = html_table_str.replace('\\', '\\\\').replace('`', '\\`').replace('$', '\\$')
        escaped_tsv = tsv_str.replace('\\', '\\\\').replace('`', '\\`').replace('$', '\\$')
        
        # 使用 HTML/JS 建立一鍵複製至剪貼簿的按鈕，支持同時寫入 HTML 與 Plain Text 格式
        copy_button_html = """
        <button id="copy-btn" style="
            background: linear-gradient(135deg, #00C6FF 0%, #0072FF 100%);
            color: white;
            border: none;
            padding: 10px 20px;
            border-radius: 8px;
            cursor: pointer;
            font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif;
            font-size: 14px;
            font-weight: 600;
            box-shadow: 0 4px 6px rgba(0,0,0,0.1);
            transition: all 0.2s ease;
            outline: none;
            display: flex;
            align-items: center;
            gap: 8px;
        ">📋 複製預覽表格 (貼上即可保留 Excel 框線與背景顏色)</button>
        
        <script>
            document.getElementById('copy-btn').addEventListener('click', function() {
                const htmlText = `__HTML_DATA__`;
                const plainText = `__TSV_DATA__`;
                
                try {
                    const blobHtml = new Blob([htmlText], { type: 'text/html' });
                    const blobText = new Blob([plainText], { type: 'text/plain' });
                    
                    const data = [new ClipboardItem({
                        'text/html': blobHtml,
                        'text/plain': blobText
                    })];
                    
                    navigator.clipboard.write(data).then(function() {
                        const btn = document.getElementById('copy-btn');
                        btn.innerHTML = '✅ 複製成功！請直接在 Excel 中按下 Ctrl+V 貼上';
                        btn.style.background = '#10B981';
                        setTimeout(() => {
                            btn.innerHTML = '📋 複製預覽表格 (貼上即可保留 Excel 框線與背景顏色)';
                            btn.style.background = 'linear-gradient(135deg, #00C6FF 0%, #0072FF 100%)';
                        }, 3000);
                    }).catch(function(err) {
                        // Fallback to plain text copy if write(ClipboardItem) fails
                        navigator.clipboard.writeText(plainText).then(function() {
                            const btn = document.getElementById('copy-btn');
                            btn.innerHTML = '⚠️ 複製文字成功（因瀏覽器限制無樣式）';
                            btn.style.background = '#F59E0B';
                            setTimeout(() => {
                                btn.innerHTML = '📋 複製預覽表格 (貼上即可保留 Excel 框線與背景顏色)';
                                btn.style.background = 'linear-gradient(135deg, #00C6FF 0%, #0072FF 100%)';
                            }, 3000);
                        });
                    });
                } catch (e) {
                    navigator.clipboard.writeText(plainText);
                }
            });
        </script>
        """.replace('__HTML_DATA__', escaped_html).replace('__TSV_DATA__', escaped_tsv)
        
        st.components.v1.html(copy_button_html, height=55)
        
        st.dataframe(
            preview_df, 
            use_container_width=True,
            hide_index=True,
            column_config={
                "Part_NO": st.column_config.TextColumn("Part_NO"),
                "Part_DESC": st.column_config.TextColumn("Model Name (Part_DESC)"),
                "Ship_NO": st.column_config.TextColumn("Ship_NO"),
                "Customer": st.column_config.TextColumn("Customer"),
                "ShippingDate": st.column_config.TextColumn("ShippingDate"),
                "SN1": st.column_config.TextColumn("SN 1"),
                "SN2": st.column_config.TextColumn("SN 2"),
                "SN3": st.column_config.TextColumn("SN 3"),
                "SN4": st.column_config.TextColumn("SN 4"),
                "SN5": st.column_config.TextColumn("SN 5")
            }
        )
        
    except Exception as e:
        st.error(f"❌ 處理過程中發生錯誤：{str(e)}")
        st.exception(e)
else:
    st.info("💡 請在上方上傳 SN Data Excel 檔案開始進行資料處理。")
